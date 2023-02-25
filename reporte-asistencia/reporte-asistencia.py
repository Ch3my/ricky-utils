# INSTALL DEPENDENCIES
# pip install openpyxl
# pip install printy

# Helpful resources
# https://stackoverflow.com/questions/42974450/iterate-over-worksheets-rows-columns

import openpyxl
from openpyxl.styles import PatternFill

from datetime import datetime
from datetime import timedelta

import os
import sys
import glob
from printy import printy

class TerminalColors:
    HEADER = "\033[95m"
    OKBLUE = "\033[94m"
    OKCYAN = "\033[96m"
    OKGREEN = "\033[92m"
    WARNING = "\033[93m"
    ERR = "\033[91m"
    ENDC = "\033[0m"
    BOLD = "\033[1m"
    UNDERLINE = "\033[4m"

class ExcelColors:
    yellow = "FFFF00"
    red = "FF0000"
    green = "00FF00"

if __name__ == "__main__":
    time_hora_cero = datetime.strptime("00:00:00", "%H:%M:%S")
    time_fin_nocturnas = datetime.strptime("08:00:00", "%H:%M:%S")
    time_inicio_nocturnas = datetime.strptime("17:00:00", "%H:%M:%S")
    # Leemos todos los excel que estan en la misma carpeta que el programa
    excelFiles = glob.glob(os.path.join(".", "*.xlsx"))

    # ██████   █████  ██ ███    ██ ████████     ██████   ██████  ██     ██ 
    # ██   ██ ██   ██ ██ ████   ██    ██        ██   ██ ██    ██ ██     ██ 
    # ██████  ███████ ██ ██ ██  ██    ██        ██████  ██    ██ ██  █  ██ 
    # ██      ██   ██ ██ ██  ██ ██    ██        ██   ██ ██    ██ ██ ███ ██ 
    # ██      ██   ██ ██ ██   ████    ██        ██   ██  ██████   ███ ███  
    # Esta funcion pinta todas las celdas de la fila que indican
    # con el color que indican
    def paint_row(sheet, row_number, color): 
        for row_cells in sheet.iter_rows(min_row=row_number, max_row=row_number):
            for cell in row_cells:
                #print('%s: cell.value=%s' % (cell, cell.value) )
                cell.fill = PatternFill("solid", fgColor=color)      


    # ███    ███  █████  ██ ███    ██
    # ████  ████ ██   ██ ██ ████   ██
    # ██ ████ ██ ███████ ██ ██ ██  ██
    # ██  ██  ██ ██   ██ ██ ██  ██ ██
    # ██      ██ ██   ██ ██ ██   ████
    while True:
        for i in range(0, len(excelFiles)):
            # Itera cada columna. Values es un array con los valores de cada columna
            # print([v.value for v in values])

            # ██████  ███████ ██████  ███████  ██████  ███    ██  █████  ███████
            # ██   ██ ██      ██   ██ ██      ██    ██ ████   ██ ██   ██ ██
            # ██████  █████   ██████  ███████ ██    ██ ██ ██  ██ ███████ ███████
            # ██      ██      ██   ██      ██ ██    ██ ██  ██ ██ ██   ██      ██
            # ██      ███████ ██   ██ ███████  ██████  ██   ████ ██   ██ ███████
            personas = []
            workbook = openpyxl.load_workbook(excelFiles[i])
            # Busca nombres de los trabajadores en hoja Especial
            sheetPersonas = workbook["Personas"]
            for row in sheetPersonas.iter_rows(min_row=2):
                personas.append(
                    {
                        "name": row[0].value,
                        "workMin": int(row[1].value),
                        "horaEntradaTeorica": row[2].value,
                    }
                )

            # ┌─┐┌┬┐┌─┐┬─┐┌┬┐  ┬ ┬┌─┐┬─┐┬┌─
            # └─┐ │ ├─┤├┬┘ │   ││││ │├┬┘├┴┐
            # └─┘ ┴ ┴ ┴┴└─ ┴   └┴┘└─┘┴└─┴ ┴
            # Selecciona Hoja por nombre
            sheet = workbook["Asistencia"]
            extrasCount = 0
            errorCount = 0
            nocturnasCount = 0

            # Iterar cada fila y recorrer todas las columnas. Comienza desde la fila 2
            for key, *values in sheet.iter_rows(min_row=2):
                # Buscamos en la lista el nombre
                thisPersona = next(
                    (item for item in personas if item["name"] == values[1].value), None
                )
                if thisPersona == None:
                    # Persona no encontrada
                    continue

                # ==== DEBUG ====
                # print("Persona encontada Fila: " + str(key.row) + ", min Trabjadso: " + str(minTrabajados))
                # print("thisPersona.workMin: " + str(thisPersona.workMin))

                # Si tiene dato vacio se lo salta
                if (
                    values[12].value == ""
                    or values[12].value == None
                    or values[13].value == ""
                    or values[13].value == None
                    or values[18].value == ""
                    or values[18].value == None
                    or values[19].value == ""
                    or values[19].value == None
                ):
                    errorCount += 1
                    paint_row(sheet, key.row, ExcelColors.red)
                    printy("Error de Informacion en fila [B]"
                           +str(key.row)+"@ (faltan datos)", "r>")
                    continue

                dateTime_hora_salida = datetime.strptime(values[19].value, "%H:%M:%S")
                dateTime_hora_entrada = datetime.strptime(values[13].value, "%H:%M:%S")

                # ┌─┐┌┐┌┌┬┐┬─┐┌─┐┌┬┐┌─┐
                # ├┤ │││ │ ├┬┘├─┤ ││├─┤
                # └─┘┘└┘ ┴ ┴└─┴ ┴─┴┘┴ ┴
                # Si entro antes de la hora teorica pero menos de 30min antes
                # contamos solo desde la hora teorica
                time_entrada_teorica = datetime.strptime(
                    thisPersona["horaEntradaTeorica"], "%H:%M"
                )
                if (
                    dateTime_hora_entrada < time_entrada_teorica
                    and dateTime_hora_entrada
                    > time_entrada_teorica - timedelta(minutes=30)
                ):
                    dateTime_entrada = datetime.strptime(
                        values[12].value + " " + thisPersona["horaEntradaTeorica"],
                        "%d/%m/%Y %H:%M",
                    )

                # Si entro despues de la hora teorica o entro antes de 30min antes de la hora de entrada
                # le contamos desde que entro
                if (
                    dateTime_hora_entrada > time_entrada_teorica
                    or dateTime_hora_entrada
                    < time_entrada_teorica - timedelta(minutes=30)
                ):
                    dateTime_entrada = datetime.strptime(
                        values[12].value + " " + values[13].value, "%d/%m/%Y %H:%M:%S"
                    )

                # ┌─┐┌─┐┬  ┬┌┬┐┌─┐
                # └─┐├─┤│  │ ││├─┤
                # └─┘┴ ┴┴─┘┴─┴┘┴ ┴
                # La hora de salida siempre se respeta
                dateTime_salida = datetime.strptime(
                    values[18].value + " " + values[19].value, "%d/%m/%Y %H:%M:%S"
                )

                minTrabajados = dateTime_salida - dateTime_entrada
                # Segundos a Minutos - Redondeamos a 2 decimales
                minTrabajados = round(minTrabajados.total_seconds() / 60, 2)

                # ██   ██ ██████  ███████     ███    ██ ██  ██████  ██   ██ ████████
                # ██   ██ ██   ██ ██          ████   ██ ██ ██       ██   ██    ██
                # ███████ ██████  ███████     ██ ██  ██ ██ ██   ███ ███████    ██
                # ██   ██ ██   ██      ██     ██  ██ ██ ██ ██    ██ ██   ██    ██
                # ██   ██ ██   ██ ███████     ██   ████ ██  ██████  ██   ██    ██
                # Horas Nocturnas, si la hora de salida es < que 08:00
                # entonces horas nocturnas = 00:00:00 - hora salida
                # Poner en columna AD (se mide en horas tambien)
                if (
                    dateTime_hora_entrada > time_inicio_nocturnas
                    and dateTime_hora_salida >= time_hora_cero
                ):
                    horasNocturnas = dateTime_hora_salida - time_hora_cero
                    # Convertimos a horas altiro, no necesitamos hacer verificaciones en minutos
                    horasNocturnas = round(
                        horasNocturnas.total_seconds() / (60 * 60), 2
                    )
                    if horasNocturnas < 10:
                        # Horas nocturnas cuentan solo hasta las 08:00
                        if dateTime_hora_salida < time_fin_nocturnas:
                            sheet[f"AD" + str(key.row)].value = horasNocturnas
                            nocturnasCount += 1
                        if dateTime_hora_salida > time_fin_nocturnas:
                            # 00:00:00 a 08:00:00 son 8 horas, el maximo
                            sheet[f"AD" + str(key.row)].value = 8
                            nocturnasCount += 1
                        paint_row(sheet, key.row, ExcelColors.yellow)

                    if horasNocturnas > 10:
                        errorCount += 1
                        paint_row(sheet, key.row, ExcelColors.red)
                        printy("Error de Informacion en fila [B]"
                               +str(key.row)+"@ (demasiadas horas nocturnas)", "r>")

                # ██   ██ ██████  ███████     ███████ ██   ██ ████████ ██████   █████
                # ██   ██ ██   ██ ██          ██       ██ ██     ██    ██   ██ ██   ██
                # ███████ ██████  ███████     █████     ███      ██    ██████  ███████
                # ██   ██ ██   ██      ██     ██       ██ ██     ██    ██   ██ ██   ██
                # ██   ██ ██   ██ ███████     ███████ ██   ██    ██    ██   ██ ██   ██
                # Si trabaja menos de 20 min extra no cuenta
                if minTrabajados < thisPersona["workMin"] + 20:
                    continue

                horasExtra = (minTrabajados - thisPersona["workMin"]) / 60
                # Si tiene mas de 5 horas extra asumiremos que es un error de informacion
                if horasExtra >= 7:
                    errorCount += 1
                    paint_row(sheet, key.row, ExcelColors.red)
                    printy("Error de Informacion en fila [B]"
                            +str(key.row)+"@ (demasiadas horas extra)", "r>")
                    continue

                sheet[f"AC" + str(key.row)].value = horasExtra
                paint_row(sheet, key.row, ExcelColors.yellow)

                extrasCount += 1
            # END for loop rows

            # Guardamos Excel e informamos resultado
            workbook.save(excelFiles[i])
            print("Archivo " + excelFiles[i] + " procesado")
            print("  Horas extras encontradas: " + str(extrasCount) + " fila(s)")
            print("  Horas nocturnas encontradas: " + str(nocturnasCount) + " fila(s)")
            print("  Filas con errores: " + str(errorCount) + " fila(s)")
        # END for loop files

        sys.exit()

